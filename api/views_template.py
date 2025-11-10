import os
from io import BytesIO
from django.http import FileResponse, HttpResponseServerError
from django.contrib.staticfiles import finders

# Nombre sugerido al descargar
DOWNLOAD_NAME = "plantilla_carga_masiva.xlsx"

# ======== Configuración de columnas ========
HEADERS = [
    "RUT",
    "Razón social",
    "Período",
    "Tipo",
    "Folio",
    "Monto",
    "Moneda",
    "Estado",
    "Observaciones",
]

TIPO_LIST = ["Factura", "Boleta", "Nota de crédito", "Otro"]
MONEDA_LIST = ["CLP", "USD", "COP", "PEN"]
ESTADO_LIST = ["Válida", "Con advertencias", "Rechazada"]

COL_WIDTHS = [15, 24, 12, 16, 12, 14, 12, 18, 28]  # aprox


def _build_xlsx_in_memory() -> BytesIO:
    """
    Genera un XLSX en memoria con:
    - Encabezados solicitados
    - 10 filas vacías
    - Validaciones para Tipo, Moneda y Estado
    - AutoFiltro en encabezados
    """
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except Exception as e:
        raise RuntimeError(
            "Falta 'openpyxl' para generar la plantilla. "
            "Instala con: pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "CargaMasiva"

    # Encabezados
    ws.append(HEADERS)

    # Estilos básicos
    fill_hdr = PatternFill("solid", fgColor="E8F5E9")  # verde claro
    font_hdr = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Anchos y estilos de encabezado
    for idx, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.fill = fill_hdr
        cell.font = font_hdr
        cell.alignment = align_center
        cell.border = border_all
        # ancho
        try:
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = COL_WIDTHS[idx-1]
        except Exception:
            pass

    # 10 filas vacías
    ROWS = 10
    COLS = len(HEADERS)
    for _ in range(ROWS):
        ws.append([""] * COLS)

    # Bordes a todo el rango de datos
    for r in range(2, ROWS + 2):
        for c in range(1, COLS + 1):
            ws.cell(row=r, column=c).border = border_all
            # Observaciones con wrap text
            if c == HEADERS.index("Observaciones") + 1:
                ws.cell(row=r, column=c).alignment = align_left

    # Validaciones (listas desplegables)
    dv_tipo = DataValidation(type="list", formula1=f'"{",".join(TIPO_LIST)}"', allow_blank=True)
    dv_moneda = DataValidation(type="list", formula1=f'"{",".join(MONEDA_LIST)}"', allow_blank=True)
    dv_estado = DataValidation(type="list", formula1=f'"{",".join(ESTADO_LIST)}"', allow_blank=True)

    ws.add_data_validation(dv_tipo)
    ws.add_data_validation(dv_moneda)
    ws.add_data_validation(dv_estado)

    # Columnas (1-based)
    col_tipo = HEADERS.index("Tipo") + 1
    col_moneda = HEADERS.index("Moneda") + 1
    col_estado = HEADERS.index("Estado") + 1

    # Rango de celdas (filas 2..ROWS+1)
    start = 2
    end = ROWS + 1
    dv_tipo.add(f"{ws.cell(row=start, column=col_tipo).coordinate}:{ws.cell(row=end, column=col_tipo).coordinate}")
    dv_moneda.add(f"{ws.cell(row=start, column=col_moneda).coordinate}:{ws.cell(row=end, column=col_moneda).coordinate}")
    dv_estado.add(f"{ws.cell(row=start, column=col_estado).coordinate}:{ws.cell(row=end, column=col_estado).coordinate}")

    # AutoFiltro en encabezados
    ws.auto_filter.ref = ws.dimensions  # aplica a todo el rango con datos

    # Congelar encabezados
    ws.freeze_panes = "A2"

    # Guardar a memoria
    mem = BytesIO()
    wb.save(mem)
    mem.seek(0)
    return mem


def calificacion_template(request):
    """
    Devuelve la plantilla XLSX con los encabezados y validaciones actualizadas.
    Se genera SIEMPRE en memoria para asegurar el formato correcto.
    """
    try:
        mem = _build_xlsx_in_memory()
        return FileResponse(mem, as_attachment=True, filename=DOWNLOAD_NAME)
    except RuntimeError as e:
        return HttpResponseServerError(str(e))
    except Exception as e:
        return HttpResponseServerError("Error generando la plantilla: %s" % e)
