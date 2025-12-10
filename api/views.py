from django.contrib.auth import get_user_model
from django.db.models import Q, F, Value
from django.db.models.functions import Replace
from django.http import HttpResponse, FileResponse
from rest_framework import status, permissions, filters, viewsets
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.generics import ListAPIView, RetrieveUpdateDestroyAPIView, CreateAPIView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.parsers import MultiPartParser, FormParser  # ⬅️ necesario para multipart
from .kafka_client import enviar_evento_calificacion


# ⬇️ NUEVO: autenticaciones para permitir cookie de sesión además de JWT
from rest_framework.authentication import SessionAuthentication
try:
    from rest_framework_simplejwt.authentication import JWTAuthentication
except Exception:
    JWTAuthentication = None  # por si no está instalado

from .models import UserFlag, Calificacion
# FxRate puede no existir en tu modelo; inténtalo opcionalmente
try:
    from .models import FxRate  # type: ignore
except Exception:  # pragma: no cover
    FxRate = None  # fallback

from .serializers import (
    UserSerializer,
    UserUpdateSerializer,
    RegisterSerializer,
    CalificacionSerializer,
)
from .permissions import (
    get_role,
    CanListUsers,
    CanCreateUsers,
    CanAssignRoles,
    CanEditBasic,
    CanDeleteUsers,
    CanResetPassword,
)

import io
import csv
import re
import os
import unicodedata
from datetime import datetime, timedelta

from django.conf import settings
from django.utils import timezone  # ⬅️ para reportes (rango por fecha)

# finder de archivos estáticos (para la plantilla XLSX exacta si se usara)
try:
    from django.contrib.staticfiles import finders
except Exception:
    finders = None  # fallback

# ⬇️ NUEVO: requests opcional para “traer de todos lados”
try:
    import requests  # type: ignore
except Exception:  # pragma: no cover
    requests = None  # si no está instalado, seguimos solo con fuentes locales

User = get_user_model()
ROLES_VALIDOS = {"Administrador", "Operador", "Auditor", "Usuario"}


# ========================= Helpers FX =========================
def fx_to_clp(amount_number: float, code: str) -> int:
    try:
        code = (code or "CLP").upper()
        if code == "CLP":
            return int(round(float(amount_number)))
        if FxRate is None:
            return int(round(float(amount_number)))
        fx = FxRate.objects.filter(code=code).first()
        if not fx:
            return int(round(float(amount_number)))
        clp = float(amount_number) * float(getattr(fx, "clp_per_unit", 1))
        return int(round(clp))
    except Exception:
        try:
            return int(round(float(amount_number or 0)))
        except Exception:
            return 0


# ========================= Utilidades varias =========================
def _ensure_role_exists(name: str):
    if name not in ROLES_VALIDOS:
        pass
    grp, _ = User.groups.rel.model.objects.get_or_create(name=name)  # Group
    return grp


def _get_flags(user: User) -> UserFlag:
    obj, _ = UserFlag.objects.get_or_create(user=user)
    return obj


def _truthy(val) -> bool:
    """Interpreta strings como 1/true/on/sí/si."""
    s = str(val or "").strip().lower()
    return s in ("1", "true", "on", "sí", "si", "yes", "y")


# ====== NUEVO: helpers para RUT flexible ======
def _clean_rut_string(s: str) -> str:
    """
    Limpia un RUT en texto, quitando puntos, guiones y espacios.
    Ej: '12.345.678-9' -> '123456789'
    """
    s = (s or "").strip()
    return re.sub(r"[.\-\s]", "", s)


def _apply_rut_filter(qs, rut_raw: str):
    """
    Aplica un filtro de RUT 'flexible', ignorando ., - y espacios
    tanto en el parámetro como en lo almacenado en BD.

    No altera cómo se guarda el RUT, solo cómo se filtra.
    """
    rut_clean = _clean_rut_string(rut_raw)
    if not rut_clean:
        return qs

    # Normalizamos el campo rut en BD quitando ., - y espacios
    qs = qs.annotate(
        rut_norm=Replace(
            Replace(
                Replace(F("rut"), Value("."), Value("")),
                Value("-"), Value("")
            ),
            Value(" "), Value("")
        )
    ).filter(rut_norm__icontains=rut_clean)
    return qs


# ================ Resolución de razón social (“de todos lados”) ================
def _resolve_razon_from_local_cache(rut: str) -> tuple[str | None, str]:
    """
    Busca en la misma tabla Calificacion algún registro previo con misma RUT
    y razón_social no vacía. Toma el más reciente.
    """
    c = (
        Calificacion.objects.filter(rut__iexact=rut)
        .exclude(Q(razon_social__isnull=True) | Q(razon_social=""))
        .order_by("-id")
        .first()
    )
    if c and c.razon_social:
        return c.razon_social.strip(), "local:calificacion"
    return None, "local:none"


def _resolve_razon_from_http(rut: str) -> tuple[str | None, str, str | None]:
    """
    Consulta servicios HTTP configurados por env:
      RESOLVE_RAZON_HTTP="https://a.example/lookup,https://b.example/find"
    Acepta JSON {"razon_social": "..."} o texto plano.
    Devuelve (razon, source, error).
    """
    if not requests:
        return None, "http:disabled", "requests-not-installed"

    urls = (os.environ.get("RESOLVE_RAZON_HTTP") or "").strip()
    if not urls:
        return None, "http:none", None

    for raw in urls.split(","):
        url = raw.strip()
        if not url:
            continue
        try:
            resp = requests.get(url, params={"rut": rut}, timeout=6)
            if resp.status_code != 200:
                continue
            razon = None
            try:
                data = resp.json()
                razon = (data.get("razon_social") or "").strip() or None
            except Exception:
                txt = (resp.text or "").strip()
                razon = txt or None
            if razon:
                return razon, f"http:{url}", None
        except Exception as e:  # pragma: no cover
            _ = e
            continue
    return None, "http:none", None


def resolve_razon_social(rut: str) -> tuple[str | None, str, str | None]:
    """
    Intenta resolver razón social en este orden:
      1) Cache local (Calificacion misma tabla)
      2) Servicios HTTP configurables por env
    Retorna (razon, source, error)
    """
    if not rut:
        return None, "invalid:empty-rut", "empty-rut"

    # 1) local
    razon, src = _resolve_razon_from_local_cache(rut)
    if razon:
        return razon, src, None

    # 2) http externos opcionales
    razon, src, err = _resolve_razon_from_http(rut)
    if razon:
        return razon, src, None

    return None, "not-found", err


# ========================= Auth y Perfil =========================
class EmailOrUsernameTokenView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        from django.contrib.auth import authenticate
        identifier = request.data.get("username") or request.data.get("email") or ""
        password = request.data.get("password") or ""
        user = None

        if "@" in identifier:
            try:
                u = User.objects.get(email__iexact=identifier.strip())
                user = authenticate(request, username=u.username, password=password)
            except User.DoesNotExist:
                user = None
        else:
            user = authenticate(request, username=identifier.strip(), password=password)

        if not user:
            return Response({"detail": "Credenciales inválidas."}, status=400)

        if not user.is_active:
            return Response(
                {"detail": "Tu cuenta está desactivada. Comunícate con un administrador."},
                status=403,
            )

        refresh = RefreshToken.for_user(user)
        flags = _get_flags(user)
        return Response(
            {
                "refresh": str(refresh),
                "access": str(refresh.access_token),
                "must_change_password": bool(flags.must_change_password),
            },
            status=200,
        )


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def me_view(request):
    data = UserSerializer(request.user).data
    data["must_change_password"] = bool(_get_flags(request.user).must_change_password)
    return Response(data)


class RegisterView(CreateAPIView):
    permission_classes = [permissions.AllowAny]
    serializer_class = RegisterSerializer


class UsersListView(ListAPIView):
    permission_classes = [permissions.IsAuthenticated, CanListUsers]
    serializer_class = UserSerializer

    def get_queryset(self):
        return User.objects.all().order_by("id")


class UsersDetailView(RetrieveUpdateDestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        if self.request.method in ("GET",):
            return [permissions.IsAuthenticated(), CanListUsers()]
        if self.request.method in ("PATCH", "PUT"):
            return [permissions.IsAuthenticated(), CanEditBasic()]
        if self.request.method == "DELETE":
            return [permissions.IsAuthenticated(), CanDeleteUsers()]
        return super().get_permissions()

    def get_serializer_class(self):
        if self.request.method in ("PATCH", "PUT"):
            return UserUpdateSerializer
        return UserSerializer

    def update(self, request, *args, **kwargs):
        from django.contrib.auth.password_validation import validate_password
        from django.core.exceptions import ValidationError as DjangoValidationError

        role = get_role(request.user)
        partial = kwargs.pop("partial", True)
        instance = self.get_object()

        before_active = bool(instance.is_active)

        allowed_for_operator = {"email", "username", "first_name", "last_name", "is_active", "profile", "phone"}

        data = dict(request.data)
        if role == "Operador":
            data = {k: v for k, v in data.items() if k in allowed_for_operator}

        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        after_active = bool(serializer.instance.is_active)
        if before_active != after_active:
            if after_active:
                print(f"[NUAMX] Cuenta reactivada: id={instance.id} email={instance.email}")
            else:
                print(f"[NUAMX] Cuenta desactivada: id={instance.id} email={instance.email}")

        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        if get_role(request.user) != "Administrador":
            return Response({"detail": "No autorizado."}, status=403)
        return super().destroy(request, *args, **kwargs)


class AssignRoleView(APIView):
    permission_classes = [permissions.IsAuthenticated, CanAssignRoles]

    def post(self, request):
        email = (request.data.get("email") or "").strip()
        role = (request.data.get("role") or "").strip()
        if not email or not role:
            return Response({"detail": "email y role son requeridos."}, status=400)
        try:
            user = User.objects.get(email__iexact=email)
        except User.DoesNotExist:
            return Response({"detail": "Usuario no existe."}, status=404)

        grp = _ensure_role_exists(role)
        user.groups.clear()
        user.groups.add(grp)
        user.save()

        return Response({"ok": True, "email": user.email, "role": grp.name}, status=200)


class UserPasswordView(APIView):
    permission_classes = [permissions.IsAuthenticated, CanResetPassword]

    def post(self, request, pk):
        from django.contrib.auth.password_validation import validate_password
        from django.core.exceptions import ValidationError as DjangoValidationError
        from django.utils import timezone

        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response({"detail": "No User matches the given query."}, status=404)

        new_pwd = request.data.get("password") or ""
        if len(new_pwd) < 8:
            return Response({"detail": "La contraseña debe tener al menos 8 caracteres."}, status=400)

        try:
            validate_password(new_pwd, user=user)
        except DjangoValidationError as e:
            return Response({"detail": " ; ".join(e.messages)}, status=400)

        role = get_role(request.user)
        force_change_req = bool(request.data.get("force_change", False))
        force_change = force_change_req if role == "Administrador" else False

        user.set_password(new_pwd)
        if hasattr(user, "last_password_change"):
            try:
                user.last_password_change = timezone.now()
            except Exception:
                pass
        user.save()

        flags = _get_flags(user)
        flags.must_change_password = force_change
        flags.save()

        return Response({"ok": True, "must_change_password": flags.must_change_password}, status=200)


class MePasswordView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        from django.contrib.auth.password_validation import validate_password
        from django.core.exceptions import ValidationError as DjangoValidationError
        from django.utils import timezone

        user = request.user
        data = request.data or {}

        new_pwd = (data.get("new_password") or (data.get("password") or "")).strip()
        current = (data.get("current_password") or "").strip()

        if not new_pwd:
            return Response({"detail": "new_password es obligatorio."}, status=400)

        flags = _get_flags(user)
        force_mode = bool(flags.must_change_password)

        if not force_mode:
            if not current:
                return Response({"detail": "current_password es obligatorio."}, status=400)
            if not user.check_password(current):
                return Response({"detail": "La contraseña actual no es válida."}, status=400)

        if len(new_pwd) < 8:
            return Response({"detail": "La nueva contraseña debe tener al menos 8 caracteres."}, status=400)

        try:
            validate_password(new_pwd, user=user)
        except DjangoValidationError as e:
            return Response({"detail": " ; ".join(e.messages)}, status=400)

        user.set_password(new_pwd)
        if hasattr(user, "last_password_change"):
            try:
                user.last_password_change = timezone.now()
            except Exception:
                pass
        user.save()

        flags.must_change_password = False
        flags.save()

        return Response({"ok": True}, status=200)


# ========================= Calificaciones (CRUD + export CSV/XLSX) =========================
class CalificacionViewSet(viewsets.ModelViewSet):
    """
    CRUD de calificaciones (auth requerida).
    """
    # ⬇️ Acepta JWT y/o cookie de sesión
    if JWTAuthentication:
        authentication_classes = (JWTAuthentication, SessionAuthentication)
    else:
        authentication_classes = (SessionAuthentication,)

    permission_classes = [permissions.IsAuthenticated]
    serializer_class = CalificacionSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ["created_at", "monto", "periodo", "rut", "folio"]
    ordering = ["-created_at"]

    def get_queryset(self):
        qs = Calificacion.objects.all()
        qp = self.request.query_params

        rut = (qp.get("rut") or "").strip()
        razon = (qp.get("razon") or "").strip()
        pdesde = (qp.get("pdesde") or "").strip()
        phasta = (qp.get("phasta") or "").strip()
        tipo = (qp.get("tipo") or "").strip()
        estado = (qp.get("estado") or "").strip()
        moneda = (qp.get("moneda") or "").strip().upper()
        # ⬇️ NUEVO: filtro de "no inscritos" (razón social vacía o NULL)
        no_inscritos = (qp.get("no_inscritos") or qp.get("noi") or "").strip().lower()
        want_noi = no_inscritos in ("1", "true", "on", "sí", "si")

        if rut:
            qs = _apply_rut_filter(qs, rut)
        if razon:
            qs = qs.filter(razon_social__icontains=razon)
        if pdesde:
            qs = qs.filter(periodo__gte=pdesde)
        if phasta:
            qs = qs.filter(periodo__lte=phasta)
        if tipo:
            qs = qs.filter(tipo_instrumento=tipo)
        if estado:
            qs = qs.filter(estado_validacion=estado)
        if moneda:
            qs = qs.filter(moneda=moneda)
        if want_noi:
            qs = qs.filter(Q(razon_social__isnull=True) | Q(razon_social=""))

        return qs

    def _build_kafka_payload(self, instance, accion: str):
        """
        Construye el payload estándar para eventos de calificación.
        """
        return {
            "accion": accion,
            "id": instance.id,
            "rut": instance.rut,
            "razon_social": instance.razon_social,
            "periodo": instance.periodo,
            "tipo_instrumento": instance.tipo_instrumento,
            "folio": instance.folio,
            "monto": int(getattr(instance, "monto", 0) or 0),
            "moneda": getattr(instance, "moneda", "CLP"),
            "estado_validacion": instance.estado_validacion,
            "created_at": instance.created_at.isoformat() if getattr(instance, "created_at", None) else None,
        }

    def perform_create(self, serializer):
        """
        Al crear una calificación, enviamos un evento a Kafka (si está habilitado).
        """
        instance = serializer.save()
        try:
            payload = self._build_kafka_payload(instance, "create")
            enviar_evento_calificacion(payload)
        except Exception:
            # Nunca rompemos el flujo de la API por un error de Kafka.
            pass

    def perform_update(self, serializer):
        """
        Al actualizar una calificación, también enviamos un evento.
        """
        instance = serializer.save()
        try:
            payload = self._build_kafka_payload(instance, "update")
            enviar_evento_calificacion(payload)
        except Exception:
            pass

    # ==================== NUEVO: Enriquecer en el listado ====================
    def list(self, request, *args, **kwargs):
        """
        Si pides ?no_inscritos=1 (o ?enrich=1), se intenta completar razon_social
        desde:
          1) cache local de Calificacion
          2) servicios HTTP en RESOLVE_RAZON_HTTP
        No persiste en BD; solo afecta la representación JSON.
        Flags:
          - enrich=1/0 (forzar encendido/apagado)
          - auto_enrich_noi=1/0 (por defecto 1) activa auto-enriquecimiento cuando no_inscritos=1
        """
        qp = request.query_params
        want_noi = _truthy(qp.get("no_inscritos") or qp.get("noi"))
        enrich_flag = qp.get("enrich")
        auto_enrich_noi = qp.get("auto_enrich_noi")

        # política: por defecto, si piden no_inscritos, enriquecemos
        auto_enrich = True if auto_enrich_noi is None else _truthy(auto_enrich_noi)
        enrich = (_truthy(enrich_flag) if enrich_flag is not None else (want_noi and auto_enrich))

        queryset = self.filter_queryset(self.get_queryset())

        if not enrich:
            # sin cambios: comportamiento 100% original
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)

        # Enriquecimiento: resolvemos por RUT y en memoria seteamos razon_social
        # para que el serializer lo devuelva completado.
        # 1) construir cache por RUT
        need = []
        for obj in queryset:
            if not (obj.razon_social or "").strip():
                if (obj.rut or "").strip():
                    need.append((obj.id, obj.rut.strip()))

        # Agrupar por RUT único
        unique_ruts = sorted({rut for _, rut in need})

        resolved_cache: dict[str, str] = {}
        for rut in unique_ruts:
            razon, _, _ = resolve_razon_social(rut)
            if razon:
                resolved_cache[rut] = razon

        # 2) aplicar sobre instancias (sin guardar)
        for obj in queryset:
            if not (obj.razon_social or "").strip():
                rut = (obj.rut or "").strip()
                if rut and rut in resolved_cache:
                    # Asignación en memoria para que el serializer lea este valor
                    obj.razon_social = resolved_cache[rut]

        # Continuar flujo DRF normal
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            # Cabecera informativa (opcional)
            resp = self.get_paginated_response(serializer.data)
            try:
                resp["X-Resolved-RUTs"] = str(len(resolved_cache))
            except Exception:
                pass
            return resp

        serializer = self.get_serializer(queryset, many=True)
        resp = Response(serializer.data)
        try:
            resp["X-Resolved-RUTs"] = str(len(resolved_cache))
        except Exception:
            pass
        return resp

    # ==================== NUEVO: Resolver "no inscritos" ====================
    @action(detail=False, methods=["post"], url_path="resolve_no_inscritos", permission_classes=[permissions.IsAuthenticated])
    def resolve_no_inscritos(self, request, *args, **kwargs):
        """
        Intenta completar razon_social para calificaciones con RUT y razon_social vacía.
        Opciones:
          - dry_run: 1 (default) => no guarda; 0 => guarda cambios
          - limit: máximo de filas a procesar (default 500)
          - overwrite: 1 para sobreescribir si existe (default 0 = solo vacíos)
          - filtros iguales a get_queryset (rut, pdesde, phasta, tipo, estado, moneda, no_inscritos)
        Fuentes:
          1) Local (otras calificaciones con mismo RUT y razon_social conocida)
          2) HTTP externos (RESOLVE_RAZON_HTTP)
        """
        qp = request.query_params
        dry_run = _truthy(qp.get("dry_run")) if qp.get("dry_run") is not None else True
        limit = int(qp.get("limit") or 500)
        overwrite = _truthy(qp.get("overwrite") or "0")

        # Tomamos el queryset respetando filtros y forzando no inscritos por defecto
        base_qs = self.get_queryset()
        if not _truthy(qp.get("no_inscritos") or qp.get("noi") or "1"):
            # si no pidieron explicitamente no_inscritos, igualmente trabajamos sobre vacíos para este endpoint
            base_qs = base_qs.filter(Q(razon_social__isnull=True) | Q(razon_social=""))

        # Solo columnas necesarias
        qs = base_qs.order_by("id").values("id", "rut", "razon_social")
        rows = list(qs[:limit])

        # Agrupamos por RUT
        by_rut = {}
        for r in rows:
            rut = (r.get("rut") or "").strip()
            if not rut:
                continue
            by_rut.setdefault(rut, []).append(r["id"])

        updated = 0
        examples = []
        errors = []

        for rut, ids in by_rut.items():
            razon, source, err = resolve_razon_social(rut)
            if err:
                errors.append({"rut": rut, "error": err, "source": source})
            if not razon:
                continue

            examples.append({"rut": rut, "razon_social": razon, "source": source, "affected": len(ids)})

            if dry_run:
                continue

            # Guardar en BD: por defecto solo vacíos, a menos que overwrite=1
            if overwrite:
                q = Calificacion.objects.filter(id__in=ids)
            else:
                q = Calificacion.objects.filter(id__in=ids).filter(Q(razon_social__isnull=True) | Q(razon_social=""))
            count = q.update(razon_social=razon)
            updated += int(count)

        return Response(
            {
                "ok": True,
                "dry_run": bool(dry_run),
                "distinct_ruts": len(by_rut),
                "processed": len(rows),
                "updated": int(updated),
                "examples": examples[:20],  # muestra algunos
                "errors": errors[:20],
            },
            status=200,
        )

    @action(detail=False, methods=["get"], url_path="export_csv")
    def export_csv(self, request, *args, **kwargs):
        # ⬇️ NUEVO: permitir enrich=1 para completar razon_social en la descarga (sin guardar en BD)
        enrich = _truthy(request.query_params.get("enrich"))
        qs = self.filter_queryset(self.get_queryset()).order_by("id")

        rows = [[
            "rut", "razon_social", "periodo", "tipo_instrumento", "folio",
            "monto", "moneda", "estado_validacion", "observaciones", "created_at"
        ]]
        for c in qs:
            razon = c.razon_social
            if enrich and not (razon or "").strip():
                resolved, _, _ = resolve_razon_social(c.rut)
                if resolved:
                    razon = resolved
            rows.append([
                c.rut,
                razon,
                c.periodo,
                c.tipo_instrumento,
                c.folio,
                str(c.monto),
                getattr(c, "moneda", "CLP"),
                c.estado_validacion,
                (c.observaciones or "").replace("\n", " ").replace("\r", " "),
                c.created_at.isoformat(),
            ])

        from io import StringIO
        sio = StringIO()
        writer = csv.writer(sio)
        for r in rows:
            writer.writerow(r)
        out = sio.getvalue()
        resp = HttpResponse(out, content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = 'attachment; filename="calificaciones.csv"'
        return resp

    @action(detail=False, methods=["post"], url_path="export_xlsx_from_rows")
    def export_xlsx_from_rows(self, request, *args, **kwargs):
        rows_in = list((request.data or {}).get("rows") or [])

        try:
            from openpyxl import Workbook
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except Exception:
            out_rows = [[
                "RUT","Razón social","Período","Tipo","Folio","Monto","Moneda","Estado","Observaciones"
            ]]
            for x in rows_in:
                out_rows.append([
                    x.get("rut") or x.get("RUT") or "",
                    x.get("razon_social") or x.get("Razón social") or "",
                    x.get("periodo") or x.get("Período") or "",
                    x.get("tipo") or x.get("Tipo") or x.get("tipo_instrumento") or "",
                    x.get("folio") or x.get("Folio") or "",
                    str(x.get("monto") or x.get("Monto") or x.get("monto_raw") or ""),
                    (x.get("moneda") or x.get("Moneda") or "").upper(),
                    x.get("estado") or x.get("Estado") or x.get("estado_validacion") or "",
                    x.get("observaciones") or x.get("Observaciones") or "",
                ])
            buf = io.StringIO()
            w = csv.writer(buf)
            for r in out_rows:
                w.writerow(r)
            resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="carga_masiva.csv"'
            return resp

        wb = Workbook()
        ws = wb.active
        ws.title = "Carga masiva"

        headers = ["RUT","Razón social","Período","Tipo","Folio","Monto","Moneda","Estado","Observaciones"]
        ws.append(headers)

        HEX_HEADER = "E8F5E9"
        HEX_ZEBRA  = "F3FBF4"
        HEX_BORDER = "DDE5DD"
        HEX_TEXT   = "0F172A"

        thin = Side(style="thin", color=HEX_BORDER)
        cell_border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for col, title in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=title)
            c.fill = PatternFill("solid", fgColor=HEX_HEADER)
            c.font = Font(bold=True, color=HEX_TEXT)
            c.alignment = Alignment(vertical="center", horizontal="center")
            c.border = cell_border

        widths = [14, 24, 12, 16, 12, 16, 12, 22, 40]
        for i, wdt in enumerate(widths, start=1):
            ws.column_dimensions[chr(64+i)].width = wdt

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:I1"
        ws.row_dimensions[1].height = 22

        # VALIDACIONES ESTRICTAS (solo opciones; sin permitir escribir otras)
        dv_tipo = DataValidation(
            type="list",
            formula1='"Factura,Boleta,Nota de crédito,Otro"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Valor inválido",
            error="Debe seleccionar una opción de la lista: Factura, Boleta, Nota de crédito u Otro.",
            errorStyle="stop",
            showInputMessage=True,
            promptTitle="Tipo",
            prompt="Seleccione: Factura, Boleta, Nota de crédito u Otro."
        )
        dv_moneda = DataValidation(
            type="list",
            formula1='"USD,COP,CLP,PEN"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Valor inválido",
            error="Debe seleccionar una moneda de la lista: USD, COP, CLP o PEN.",
            errorStyle="stop",
            showInputMessage=True,
            promptTitle="Moneda",
            prompt="Seleccione: USD, COP, CLP o PEN."
        )
        dv_estado = DataValidation(
            type="list",
            formula1='"Válida,Con advertencias,Rechazada"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Valor inválido",
            error="Debe seleccionar una opción: Válida, Con advertencias o Rechazada.",
            errorStyle="stop",
            showInputMessage=True,
            promptTitle="Estado",
            prompt="Seleccione: Válida, Con advertencias o Rechazada."
        )
        ws.add_data_validation(dv_tipo)
        ws.add_data_validation(dv_moneda)
        ws.add_data_validation(dv_estado)

        def parse_amount(moneda: str, monto_texto: str):
            m = (moneda or "").upper()
            s = str(monto_texto or "").strip()
            if m in ("USD", "PEN"):
                try:
                    return float(s.replace(".", "").replace(",", ".")) if s else None
                except Exception:
                    return None
            else:
                try:
                    num = re.sub(r"[^\d]", "", s)
                    return int(num) if num else None
                except Exception:
                    return None

        start = 2
        for i, x in enumerate(rows_in, start=start):
            rut  = x.get("rut") or x.get("RUT") or ""
            raz  = x.get("razon_social") or x.get("Razón social") or ""
            per  = x.get("periodo") or x.get("Período") or ""
            tip  = x.get("tipo") or x.get("Tipo") or x.get("tipo_instrumento") or ""
            fol  = x.get("folio") or x.get("Folio") or ""
            mon_txt = x.get("moneda") or x.get("Moneda") or ""
            mon = (mon_txt or "").upper()
            est  = x.get("estado") or x.get("Estado") or x.get("estado_validacion") or ""
            obs  = x.get("observaciones") or x.get("Observaciones") or ""
            monto_txt = x.get("monto") or x.get("Monto") or x.get("monto_raw") or ""

            ws.append([rut, raz, per, tip, fol, None, mon, est, obs])

            if i % 2 == 0:
                for col in range(1, 10):
                    ws.cell(i, col).fill = PatternFill("solid", fgColor=HEX_ZEBRA)

            for col in range(1, 10):
                cell = ws.cell(i, col)
                cell.border = cell_border
                cell.font = Font(color=HEX_TEXT)
                if col in (1, 3, 4, 5, 7, 8):
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                elif col == 6:
                    cell.alignment = Alignment(vertical="center", horizontal="right")
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Aplicar validaciones celda por celda
            dv_tipo.add(ws.cell(i, 4))
            dv_moneda.add(ws.cell(i, 7))
            dv_estado.add(ws.cell(i, 8))

            num = parse_amount(mon, monto_txt)
            c_monto = ws.cell(i, 6, num)
            if num is None:
                c_monto.value = None
            c_monto.number_format = "#,##0.00" if mon in ("USD", "PEN") else "#,##0"

            ws.row_dimensions[i].height = 18

        last_row = max(ws.max_row, 1)
        from openpyxl.worksheet.table import Table, TableStyleInfo
        table_ref = f"A1:I{last_row}"
        table = Table(displayName="Calificaciones", ref=table_ref)
        table_style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = table_style
        ws.add_table(table)

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="carga_masiva.xlsx"'
        return resp

    # ====== Endpoint de estadísticas para el dashboard (con auth múltiple) ======
    @action(detail=False, methods=["get"], url_path="stats", permission_classes=[permissions.IsAuthenticated])
    def stats(self, request, *args, **kwargs):
        from collections import Counter, defaultdict

        qs = Calificacion.objects.all()

        total = qs.count()

        hoy = timezone.now().date()
        inicio = hoy - timedelta(days=13)
        por_dia = defaultdict(int)
        for c in qs.filter(created_at__date__gte=inicio, created_at__date__lte=hoy):
            por_dia[c.created_at.date()] += 1
        serie_14d = []
        for i in range(14):
            d = inicio + timedelta(days=i)
            serie_14d.append({"date": d.isoformat(), "count": por_dia[d]})

        estados_cnt = Counter(qs.values_list("estado_validacion", flat=True))
        tipos_cnt   = Counter(qs.values_list("tipo_instrumento", flat=True))

        errores_14d = qs.filter(
            created_at__date__gte=inicio,
            created_at__date__lte=hoy,
            estado_validacion__in=["Con advertencias", "Rechazada"],
        ).count()

        data = {
            "total_registros": total,
            "registros_por_dia": serie_14d,
            "estados": {
                "Válida": int(estados_cnt.get("Válida", 0)),
                "Con advertencias": int(estados_cnt.get("Con advertencias", 0)),
                "Rechazada": int(estados_cnt.get("Rechazada", 0)),
            },
            "tipos": {
                "Factura": int(tipos_cnt.get("Factura", 0)),
                "Boleta": int(tipos_cnt.get("Boleta", 0)),
                "Nota de crédito": int(tipos_cnt.get("Nota de crédito", 0)),
                "Otro": int(tipos_cnt.get("Otro", 0)),
            },
            "errores_ultimos_14d": int(errores_14d),
        }
        return Response(data, status=200)


# ========================= Carga masiva: plantilla/preview/commit =========================
class CalificacionTemplateView(APIView):
    """
    Genera y entrega un XLSX con EXACTAMENTE las mismas opciones de Carga Masiva:
    - Columnas: RUT, Razón social, Período, Tipo, Folio, Monto, Moneda, Estado, Observaciones
    - Listas (solo selección, sin escritura libre):
        * Tipo:    Factura, Boleta, Nota de crédito, Otro
        * Moneda:  USD, COP, CLP, PEN
        * Estado:  Válida, Con advertencias, Rechazada
    - 10 filas vacías, cabecera verde suave, zebra, autofiltro y panes congelados.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except Exception:
            return Response(
                {"detail": "Servidor sin 'openpyxl'. Instálalo: pip install openpyxl"},
                status=500,
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Carga masiva"

        headers = ["RUT","Razón social","Período","Tipo","Folio","Monto","Moneda","Estado","Observaciones"]
        ws.append(headers)

        HEX_HEADER = "E8F5E9"
        HEX_ZEBRA  = "F3FBF4"
        HEX_BORDER = "DDE5DD"
        HEX_TEXT   = "0F172A"

        thin = Side(style="thin", color=HEX_BORDER)
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for col, title in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=title)
            c.fill = PatternFill("solid", fgColor=HEX_HEADER)
            c.font = Font(bold=True, color=HEX_TEXT)
            c.alignment = Alignment(vertical="center", horizontal="center")
            c.border = border

        widths = [14, 26, 12, 18, 12, 16, 12, 22, 40]
        for i, wdt in enumerate(widths, start=1):
            ws.column_dimensions[chr(64+i)].width = wdt

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:I1"
        ws.row_dimensions[1].height = 22

        # VALIDACIONES ESTRICTAS (solo opciones; sin escritura libre)
        dv_tipo = DataValidation(
            type="list",
            formula1='"Factura,Boleta,Nota de crédito,Otro"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Valor inválido",
            error="Debe seleccionar una opción de la lista: Factura, Boleta, Nota de crédito u Otro.",
            errorStyle="stop",
            showInputMessage=True,
            promptTitle="Tipo",
            prompt="Seleccione: Factura, Boleta, Nota de crédito u Otro."
        )
        dv_moneda = DataValidation(
            type="list",
            formula1='"USD,COP,CLP,PEN"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Valor inválido",
            error="Debe seleccionar una moneda de la lista: USD, COP, CLP o PEN.",
            errorStyle="stop",
            showInputMessage=True,
            promptTitle="Moneda",
            prompt="Seleccione: USD, COP, CLP o PEN."
        )
        dv_estado = DataValidation(
            type="list",
            formula1='"Válida,Con advertencias,Rechazada"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Valor inválido",
            error="Debe seleccionar una opción: Válida, Con advertencias o Rechazada.",
            errorStyle="stop",
            showInputMessage=True,
            promptTitle="Estado",
            prompt="Seleccione: Válida, Con advertencias o Rechazada."
        )
        ws.add_data_validation(dv_tipo)
        ws.add_data_validation(dv_moneda)
        ws.add_data_validation(dv_estado)

        # 10 filas vacías con validaciones
        for i in range(2, 12):
            ws.append(["", "", "", "", "", None, "", "", ""])
            if i % 2 == 0:
                for col in range(1, 10):
                    ws.cell(i, col).fill = PatternFill("solid", fgColor=HEX_ZEBRA)
            for col in range(1, 10):
                cell = ws.cell(i, col)
                cell.border = border
                cell.font = Font(color=HEX_TEXT)
                if col in (1, 3, 4, 5, 7, 8):
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                elif col == 6:
                    cell.alignment = Alignment(vertical="center", horizontal="right")
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            dv_tipo.add(f"D{i}")
            dv_moneda.add(f"G{i}")
            dv_estado.add(f"H{i}")
            ws.cell(i, 6).number_format = "#,##0"  # entero por defecto

        # Tabla
        last_row = ws.max_row
        table = Table(displayName="Calificaciones", ref=f"A1:I{last_row}")
        table_style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        table.tableStyleInfo = table_style
        ws.add_table(table)

        bio = io.BytesIO()
        wb.save(bio); bio.seek(0)
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="plantilla_carga_masiva.xlsx"'
        return resp


class CalificacionBulkPreviewView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def _norm(self, s: str) -> str:
        s = (s or "").strip().lower()
        s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
        s = s.replace("razon social", "razon_social")
        s = s.replace("período", "periodo")
        s = s.replace("tipo de documento", "tipo")
        s = s.replace("tipo documento", "tipo")
        s = s.replace("nota de credito", "nota de crédito")
        return s

    def _parse_rows(self, request):
        if isinstance(request.data, dict) and "rows" in request.data:
            raw_rows = list(request.data.get("rows") or [])
            parsed = []
            for x in raw_rows:
                parsed.append({
                    "rut":               (x.get("rut") or x.get("RUT") or "").strip(),
                    "razon_social":      (x.get("razon_social") or x.get("Razón social") or "").strip(),
                    "periodo":           (x.get("periodo") or x.get("Período") or "").strip(),
                    "tipo_instrumento":  (x.get("tipo_instrumento") or x.get("tipo") or x.get("Tipo") or "Factura").strip(),
                    "folio":             (x.get("folio") or x.get("Folio") or "").strip(),
                    "monto_raw":         str(x.get("monto_raw") or x.get("monto") or x.get("Monto") or "").replace(".", ",").strip(),
                    "moneda":            (x.get("moneda") or x.get("Moneda") or "CLP").strip().upper(),
                    "estado_validacion": (x.get("estado_validacion") or x.get("estado") or x.get("Estado") or "Válida").strip(),
                    "observaciones":     (x.get("observaciones") or x.get("Observaciones") or "").strip(),
                })
            return parsed, None

        up = request.FILES.get("file")
        if not up:
            return [], "No llegó archivo 'file' en el form-data."

        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            return [], "Servidor sin 'openpyxl'. Instálalo o envía JSON."

        wb = load_workbook(up, data_only=True, read_only=True)
        ws = wb.active

        headers = []
        for c in ws[1]:
            headers.append(self._norm(str(c.value or "")))

        idx = {h: i for i, h in enumerate(headers)}

        def col(*names):
            for n in names:
                n2 = self._norm(n)
                if n2 in idx:
                    return idx[n2]
            return None

        need = {
            "rut": col("RUT", "rut"),
            "razon_social": col("Razón social", "razon_social"),
            "periodo": col("Período", "periodo"),
            "tipo": col("Tipo", "tipo", "tipo de documento", "tipo documento"),
            "folio": col("Folio", "folio"),
            "monto": col("Monto", "monto"),
            "moneda": col("Moneda", "moneda"),
            "estado": col("Estado", "estado"),
            "observaciones": col("Observaciones", "observaciones", "obs"),
        }

        required = ("rut", "razon_social", "periodo", "tipo", "folio", "monto", "moneda", "estado")
        missing = [k for k in required if need[k] is None]
        if missing:
            return [], ("Encabezados faltantes o distintos: " + ", ".join(missing))

        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            vals = list(r or [])
            if not any(str(v).strip() for v in vals if v is not None):
                continue

            def gv(key):
                i = need[key]
                return "" if i is None or i >= len(vals) else ("" if vals[i] is None else str(vals[i]))

            periodo_cell = vals[need["periodo"]] if need["periodo"] is not None and need["periodo"] < len(vals) else ""
            if isinstance(periodo_cell, datetime):
                per_txt = periodo_cell.strftime("%Y-%m")
            else:
                per_raw = str(periodo_cell or "").strip()
                if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", per_raw):
                    try:
                        dt = datetime.strptime(per_raw, "%d/%m/%Y")
                        per_txt = dt.strftime("%Y-%m")
                    except Exception:
                        per_txt = per_raw
                else:
                    per_txt = per_raw

            rows.append({
                "rut":               gv("rut").strip(),
                "razon_social":      gv("razon_social").strip(),
                "periodo":           per_txt,
                "tipo_instrumento":  (gv("tipo") or "Factura").strip(),
                "folio":             gv("folio").strip(),
                "monto_raw":         str(gv("monto")).replace(".", ",").strip(),
                "moneda":            (gv("moneda").strip().upper() or "CLP"),
                "estado_validacion": (gv("estado") or "Válida").strip(),
                "observaciones":     gv("observaciones").strip(),
            })

        return rows, None

    def _validate_one(self, row):
        errs = []

        rut = (row.get("rut") or "").strip()
        if not rut:
            errs.append("RUT es obligatorio.")

        periodo = (row.get("periodo") or "").strip()
        if not re.match(r"^\d{4}-\d{2}$", periodo):
            errs.append("Período debe ser YYYY-MM.")

        tipo = (row.get("tipo_instrumento") or "").strip()
        if not tipo:
            errs.append("Tipo es obligatorio.")

        folio = (row.get("folio") or "").strip()
        if not folio:
            errs.append("Folio es obligatorio.")

        moneda = (row.get("moneda") or "CLP").upper()
        if moneda not in ("CLP","USD","COP","PEN"):
            errs.append("Moneda inválida (use CLP, USD, COP o PEN).")

        monto_raw = str(row.get("monto_raw") or "")
        if moneda in ("USD","PEN"):
            if not re.match(r"^\d+(,\d{1,2})?$", monto_raw or ""):
                errs.append("Monto debe usar coma decimal (ej: 123,45) para USD/PEN.")
            monto_number = float((monto_raw or "0").replace(",", ".") or 0)
        else:
            if not re.match(r"^\d+$", re.sub(r"[^\d]", "", monto_raw or "0")):
                errs.append("Monto debe ser entero (sin decimales) para CLP/COP.")
            monto_number = float(re.sub(r"[^\d]", "", monto_raw or "0") or 0)

        estado = (row.get("estado_validacion") or "").strip()
        if estado not in ("Válida","Con advertencias","Rechazada"):
            errs.append("Estado inválido.")

        monto_clp = fx_to_clp(monto_number, moneda)

        out = dict(row)
        out["errors"] = errs
        out["monto_number"] = monto_number
        out["monto_clp"] = monto_clp
        return out

    def post(self, request):
        rows, import_err = self._parse_rows(request)
        if import_err:
            return Response({"currency": "MIXED", "rows": [], "errors": [import_err]}, status=400)

        out_rows = [self._validate_one(r) for r in rows]
        if not out_rows:
            return Response({"currency": "MIXED", "rows": [], "errors": ["El archivo no contiene filas válidas."]}, status=400)

        monedas = {(r.get("moneda") or "CLP").upper() for r in rows}
        summary_currency = list(monedas)[0] if len(monedas) == 1 else "MIXED"

        return Response({"currency": summary_currency, "rows": out_rows, "errors": []}, status=200)


class CalificacionBulkCommitView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        default_currency = (request.data.get("currency") or "CLP").upper()
        raw_rows = list(request.data.get("rows") or [])
        if not raw_rows:
            return Response({"detail": "rows es obligatorio."}, status=400)

        created = 0
        ids = []

        for x in raw_rows:
            r = {
                "rut":            (x.get("rut") or x.get("RUT") or "").strip(),
                "razon_social":   (x.get("razon_social") or x.get("Razón social") or "").strip(),
                "periodo":        (x.get("periodo") or x.get("Período") or "").strip(),
                "tipo_instrumento": (x.get("tipo_instrumento") or x.get("tipo") or x.get("Tipo") or "Factura").strip(),
                "folio":          (x.get("folio") or x.get("Folio") or "").strip(),
                "monto_raw":      str(x.get("monto_raw") or x.get("monto") or x.get("Monto") or "").strip(),
                "moneda":         (x.get("moneda") or x.get("Moneda") or default_currency).strip().upper(),
                "estado_validacion": (x.get("estado_validacion") or x.get("estado") or x.get("Estado") or "Válida").strip(),
                "observaciones":  (x.get("observaciones") or x.get("Observaciones") or "").strip(),
            }

            per = r["periodo"]
            if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", per):
                try:
                    dt = datetime.strptime(per, "%d/%m/%Y")
                    r["periodo"] = dt.strftime("%Y-%m")
                except Exception:
                    pass

            moneda = r["moneda"] or default_currency

            if moneda in ("USD", "PEN"):
                monto_number = float((r["monto_raw"] or "0").replace(",", ".") or 0)
            else:
                monto_number = float(re.sub(r"[^\d]", "", r["monto_raw"] or "0") or 0)

            monto_save = int(round(monto_number))

            obj = Calificacion.objects.create(
                rut=r["rut"],
                razon_social=r["razon_social"],
                periodo=r["periodo"],
                tipo_instrumento=r["tipo_instrumento"],
                folio=r["folio"],
                monto=monto_save,
                moneda=moneda,
                estado_validacion=r["estado_validacion"],
                observaciones=r["observaciones"],
            )
            created += 1
            ids.append(obj.id)

        return Response({"ok": True, "created": created, "ids": ids}, status=200)


# ========================= Reportes (descarga XLSX/CSV con formato bonito) =========================
class ReporteExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    TITLES = {
        "diario":  "Reporte diario",
        "semanal": "Reporte semanal",
        "mensual": "Reporte mensual",
    }

    def _rango(self, scope: str):
        now = timezone.now()
        if scope == "diario":
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif scope == "semanal":
            start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, now

    def _filtrar(self, request):
        qs = Calificacion.objects.all()
        qp = request.query_params
        rut     = (qp.get("rut") or "").strip()
        razon   = (qp.get("razon") or "").strip()
        pdesde  = (qp.get("pdesde") or "").strip()
        phasta  = (qp.get("phasta") or "").strip()
        tipo    = (qp.get("tipo") or "").strip()
        estado  = (qp.get("estado") or "").strip()
        moneda  = (qp.get("moneda") or "").strip().upper()
        # ⬇️ NUEVO: soporte para no_inscritos también en reportes
        no_inscritos = (qp.get("no_inscritos") or qp.get("noi") or "").strip().lower()
        want_noi = no_inscritos in ("1", "true", "on", "sí", "si")

        if rut:
            qs = _apply_rut_filter(qs, rut)
        if razon:
            qs = qs.filter(razon_social__icontains=razon)
        if pdesde:
            qs = qs.filter(periodo__gte=pdesde)
        if phasta:
            qs = qs.filter(periodo__lte=phasta)
        if tipo:
            qs = qs.filter(tipo_instrumento=tipo)
        if estado:
            qs = qs.filter(estado_validacion=estado)
        if moneda:
            qs = qs.filter(moneda=moneda)
        if want_noi:
            qs = qs.filter(Q(razon_social__isnull=True) | Q(razon_social=""))
        return qs

    def get(self, request):
        scope = (request.query_params.get("scope") or "mensual").lower()
        fmt   = (request.query_params.get("format") or "xlsx").lower()
        enrich = _truthy(request.query_params.get("enrich"))  # ⬅️ NUEVO
        if scope not in ("diario", "semanal", "mensual"):
            return Response({"detail": "scope inválido"}, status=400)

        title = self.TITLES.get(scope, "Reporte")
        start, end = self._rango(scope)
        qs = self._filtrar(request).filter(created_at__gte=start, created_at__lte=end).order_by("id")

        headers = [
            "RUT","Razón social","Período","Tipo","Folio",
            "Monto","Moneda","Estado","Observaciones","Creado en"
        ]
        data = []
        tz = timezone.get_current_timezone()
        for c in qs:
            razon = c.razon_social
            if enrich and not (razon or "").strip():
                resolved, _, _ = resolve_razon_social(c.rut)
                if resolved:
                    razon = resolved
            data.append([
                c.rut,
                razon,
                c.periodo,
                c.tipo_instrumento,
                c.folio,
                int(c.monto),
                getattr(c, "moneda", "CLP"),
                c.estado_validacion,
                (c.observaciones or "").replace("\n", " ").replace("\r", " "),
                c.created_at.astimezone(tz).strftime("%Y-%m-%d %H:%M:%S"),
            ])

        today_str = datetime.now().strftime("%Y-%m-%d")
        base_name = f"{title.replace(' ', '_').lower()}_{today_str}"

        if fmt == "csv":
            sio = io.StringIO()
            w = csv.writer(sio)
            w.writerow([title]); w.writerow([]); w.writerow(headers)
            for row in data:
                w.writerow(row)
            resp = HttpResponse(sio.getvalue(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{base_name}.csv"'
            return resp

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except Exception:
            sio = io.StringIO()
            w = csv.writer(sio)
            w.writerow([title]); w.writerow([]); w.writerow(headers)
            for row in data:
                w.writerow(row)
            resp = HttpResponse(sio.getvalue(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{base_name}.csv"'
            return resp

        wb = Workbook()

        HEX_TITLE  = "CDEFD6"
        HEX_HEADER = "DCFCE7"
        HEX_ZEBRA  = "F0FDF4"
        HEX_BORDER = "E5E7EB"
        HEX_TEXT   = "0F172A"
        thin = Side(style="thin", color=HEX_BORDER)
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        ws = wb.active
        ws.title = title

        ws.merge_cells("A1:J1")
        c_title = ws["A1"]
        c_title.value = title
        c_title.fill = PatternFill("solid", fgColor=HEX_TITLE)
        c_title.font = Font(size=14, bold=True, color=HEX_TEXT)
        c_title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.append([""] * len(headers))
        ws.row_dimensions[2].height = 6

        ws.append(headers)
        for col in range(1, len(headers)+1):
            c = ws.cell(3, col)
            c.font = Font(bold=True, color=HEX_TEXT)
            c.fill = PatternFill("solid", fgColor=HEX_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        ws.row_dimensions[3].height = 22
        ws.freeze_panes = "A4"

        start_row = 4
        for i, row in enumerate(data, start=start_row):
            ws.append(row)
            if i % 2 == 0:
                for col in range(1, len(headers)+1):
                    ws.cell(i, col).fill = PatternFill("solid", fgColor=HEX_ZEBRA)
            for col in range(1, len(headers)+1):
                cell = ws.cell(i, col)
                cell.border = border
                cell.font = Font(color=HEX_TEXT)
                if col in (1,3,4,5,7,8,10):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col == 6:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(i, 6).number_format = "#,##0"

        last_row = ws.max_row
        last_col = ws.max_column
        ref = f"A3:{chr(64+last_col)}{last_row}"
        table = Table(displayName="TablaReporte", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight21", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)

        widths = [14, 30, 10, 16, 12, 16, 10, 18, 40, 19]
        for idx, wdt in enumerate(widths, start=1):
            ws.column_dimensions[chr(64+idx)].width = wdt

        total_row = last_row + 2
        ws.cell(total_row, 1, "Totales").font = Font(bold=True)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
        ws.cell(total_row, 6, f"=SUBTOTAL(9,F4:F{last_row})").number_format = "#,##0"
        ws.cell(total_row, 9, "Documentos:").alignment = Alignment(horizontal="right")
        ws.cell(total_row, 10, f"=SUBTOTAL(3,A4:A{last_row})")

        from collections import defaultdict
        res = defaultdict(lambda: {"count": 0, "sum": 0})
        for r in data:
            mon = r[6] or "CLP"
            res[mon]["count"] += 1
            res[mon]["sum"] += int(r[5] or 0)

        ws2 = wb.create_sheet("Resumen")
        ws2.append(["Moneda", "Documentos", "Suma monto"])
        for col in range(1, 3+1):
            c = ws2.cell(1, col)
            c.font = Font(bold=True, color=HEX_TEXT)
            c.fill = PatternFill("solid", fgColor=HEX_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        for mon, agg in res.items():
            ws2.append([mon, agg["count"], agg["sum"]])
        for r in range(2, ws2.max_row+1):
            ws2.cell(r, 3).number_format = "#,##0"
            for c in range(1, 4):
                ws2.cell(r, c).border = border
        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 14
        ws2.column_dimensions["C"].width = 16

        info_row = ws2.max_row + 2
        tz = timezone.get_current_timezone()
        ws2.cell(info_row, 1, f"Generado: {timezone.now().astimezone(tz).strftime('%Y-%m-%d %H:%M:%S')}")
        ws2.cell(info_row+1, 1, f"Alcance: {scope.title()}")
        ws2.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=3)
        ws2.merge_cells(start_row=info_row+1, start_column=1, end_row=info_row+1, end_column=3)

        bio = io.BytesIO()
        wb.save(bio); bio.seek(0)
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{base_name}.xlsx"'
        return resp
